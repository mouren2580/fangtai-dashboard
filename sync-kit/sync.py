# -*- coding: utf-8 -*-
"""
方太西北大区 服务产品达成看板 —— 一键同步脚本（家庭电脑可用，零外部技能依赖）

把「最新月报 Excel」→ 刷新 dashboard.html 里的主数据 / 服务工程师 / 延保 / 周数据
→（可选）发布到 GitHub Pages，一条龙完成。

设计目标：可移植。所有路径用参数或相对路径，不再写死 E:/ 或 D:/ 或 .workbuddy 技能目录。
唯一依赖：本地同目录的 build_dashboard.py（已随本包提供） + openpyxl。

用法（家里电脑）:
  # 1) 只刷新本地 dashboard.html，不发布（先验证）
  python sync.py --excel "2026年8月西北服务产品.xlsx" --date 2026-08-16

  # 2) 刷新并发布到线上（需要 GitHub 个人令牌）
  python sync.py --excel "2026年8月西北服务产品.xlsx" --date 2026-08-16 --token ghp_xxxxxxxx
  # 或把令牌放 .gh_token 文件里：
  python sync.py --excel "2026年8月西北服务产品.xlsx" --date 2026-08-16 --token-file .gh_token

  # 3) 不推进「数据更新时间」戳（默认就是这样，符合铁律：用户未确认更新不滚动链接日期）
  #    只有确实拿到新数据并要对外更新时，加 --bump-stamp

  # 4) 指定输出文件（默认覆盖同目录 dashboard.html）
  python sync.py --excel "x.xlsx" --date 2026-08-16 --html dashboard.html
"""
import sys, os, re, json, argparse, datetime, base64, urllib.request, urllib.error, ssl, time
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from build_dashboard import build  # 本地副本，无需安装技能

ITEMS = ["清洗保养", "延保服务", "清洁耗材", "止回阀", "装饰罩", "水/气路金属软管",
         "排烟出风管", "水气路通用辅材", "其他产品线专用", "其他水路辅材", "其他气路辅材",
         "橱柜类辅材", "燃气管快速接头", "工具", "电路辅材"]


# ---------------- 通用工具 ----------------
def _fix(s):
    if isinstance(s, str):
        try:
            return s.encode("latin-1").decode("gbk")
        except Exception:
            return s
    return s


def _num(v):
    try:
        return round(float(v), 2)
    except Exception:
        return 0.0


def replace_var(html, varname, obj):
    """替换 dashboard.html 里形如 `let VAR = {...};`（可能带行尾注释）的单行声明。"""
    pat = re.compile(r"^(?:const|let)\s+" + re.escape(varname) + r" = .*$", re.M)
    m = pat.search(html)
    if not m:
        raise SystemExit("未找到 %s 单行声明，请确认 dashboard.html 未被破坏" % varname)
    new = "let %s = %s;" % (varname, json.dumps(obj, ensure_ascii=False, separators=(",", ":")))
    return html[:m.start()] + new + html[m.end():]


# ---------------- 服务工程师（TECH_DATA）----------------
def build_tech(src):
    import openpyxl
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["服务产品收入统计"]
    techs, source_split = {}, defaultdict(float)
    full_total = 0.0
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 25:
            continue
        amount = _num(r[24])
        full_total += amount
        src_name = _fix(r[8]) if r[8] else ""
        source_split[src_name] += amount
        code = _fix(r[7]) if r[7] else ""
        if not code:
            continue  # 无工程师编码 = 网点买断，不计入工程师排行
        t = techs.get(code)
        if t is None:
            t = {"id": code, "name": _fix(r[6]) or "", "office": _fix(r[2]) or "",
                 "center": _fix(r[1]) or "", "nets": set(), "items": [0.0] * 15, "amount": 0.0}
            techs[code] = t
        t["amount"] += amount
        for i, ci in enumerate(range(9, 24)):
            t["items"][i] += _num(r[ci])
        if r[3]:
            t["nets"].add(_fix(r[3]))
        if not t["office"] and r[2]:
            t["office"] = _fix(r[2])
        if not t["center"] and r[1]:
            t["center"] = _fix(r[1])
    tech_list, offices = [], set()
    for t in techs.values():
        if t["office"]:
            offices.add(t["office"])
        tech_list.append({
            "id": t["id"], "name": t["name"], "office": t["office"], "center": t["center"],
            "net": "、".join(sorted(t["nets"])) if t["nets"] else "",
            "cnt": sum(1 for v in t["items"] if v > 0),
            "amount": round(t["amount"], 2),
            "items": {ITEMS[i]: round(v, 2) for i, v in enumerate(t["items"])}})
    tech_list.sort(key=lambda x: x["amount"], reverse=True)
    total = round(sum(t["amount"] for t in tech_list), 2)
    full_total = round(full_total, 2)
    other = round(full_total - total, 2)
    tech = {"techs": tech_list, "offices": sorted(offices), "items": ITEMS,
            "total": total, "count": len(tech_list),
            "meta": {"口径": "服务产品收入统计·按服务工程师(姓名+编码)汇总『合计』(工单消耗)",
                     "来源": "服务产品收入统计", "截止": "报表全量",
                     "报表总合计": full_total, "其他未归属": other,
                     "工程师工单消耗": total, "网点买断": other}}
    return tech, dict(source_split)


# ---------------- 延保（EXTEND_WARRANTY_DATA）----------------
def gen_extend(excel, cutoff):
    import openpyxl
    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    target = next((s for s in wb.sheetnames if ("服务项目" in s) or ("CSM" in s.upper())), None)
    ws = wb[target]
    KW = ["销售分类", "服务收费项目", "服务工程师", "服务工程师编码", "上缴金额", "办事处"]
    hdr_row = hdr = None
    for ri in range(1, 7):
        row = list(next(ws.iter_rows(min_row=ri, max_row=ri, values_only=True)))
        flat = [str(x) for x in row if x is not None]
        if any(k in f for f in flat for k in KW):
            hdr_row, hdr = ri, row
            break
    idx = {}
    for ci, v in enumerate(hdr):
        if v is None:
            continue
        s = str(v).strip()
        if s in KW:
            idx[s] = ci
    assert all(k in idx for k in ["销售分类", "服务收费项目", "服务工程师", "上缴金额"]), idx
    c_cat, c_item, c_tech = idx["销售分类"], idx["服务收费项目"], idx["服务工程师"]
    c_code, c_pay, c_off = idx.get("服务工程师编码"), idx["上缴金额"], idx.get("办事处")
    maxc = max(x for x in (c_cat, c_item, c_tech, c_pay, c_code, c_off) if x is not None)

    def yr_of(s):
        s = str(s)
        if "一年" in s or "1年" in s:
            return "y1"
        if "两年" in s or "二年" in s or "2年" in s:
            return "y2"
        if "三年" in s or "3年" in s:
            return "y3"
        return None

    techs, offices = {}, set()
    tot = {"y1": {"qty": 0, "amt": 0.0}, "y2": {"qty": 0, "amt": 0.0}, "y3": {"qty": 0, "amt": 0.0}}
    n = 0
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row or len(row) <= maxc:
            continue
        if str(row[c_cat]).strip() != "延保服务":
            continue
        yk = yr_of(row[c_item])
        if yk is None:
            continue
        name = str(row[c_tech]).strip() if row[c_tech] else "（未知）"
        code = str(row[c_code]).strip() if (c_code is not None and row[c_code]) else ""
        office = str(row[c_off]).strip() if (c_off is not None and row[c_off]) else ""
        if office:
            offices.add(office)
        key = (name, code)
        t = techs.get(key)
        if t is None:
            t = {"id": (code or name), "name": name, "office": office,
                 "y1": {"qty": 0, "amt": 0.0}, "y2": {"qty": 0, "amt": 0.0}, "y3": {"qty": 0, "amt": 0.0}}
            techs[key] = t
        elif office and not t["office"]:
            t["office"] = office
        amt = _num(row[c_pay])
        t[yk]["qty"] += 1
        t[yk]["amt"] += amt
        tot[yk]["qty"] += 1
        tot[yk]["amt"] += amt
        n += 1
    tech_list = list(techs.values())
    for t in tech_list:
        t["tqty"] = t["y1"]["qty"] + t["y2"]["qty"] + t["y3"]["qty"]
        t["tamt"] = round(t["y1"]["amt"] + t["y2"]["amt"] + t["y3"]["amt"], 2)
    tech_list.sort(key=lambda x: x["tamt"], reverse=True)
    return {"meta": {
                "来源": "CSM服务项目", "筛选": "销售分类=延保服务", "金额口径": "上缴金额",
                "年限区分": "服务收费项目列(含一年/两年/三年)",
                "截止": cutoff.strftime("%Y-%m-%d"), "工程师数": len(tech_list)},
            "offices": sorted(offices), "techs": tech_list, "total": tot}


# ---------------- 周数据（WEEKLY_DATA）----------------
def _svc_start(month, year):
    if month == 1:
        return datetime.date(year - 1, 12, 28)
    return datetime.date(year, month - 1, 28)


def gen_weekly(excel, cutoff):
    import openpyxl
    wb = openpyxl.load_workbook(excel, data_only=True, read_only=True)

    def find_sheet(p):
        for ws in wb.worksheets:
            if p in ws.title:
                return ws

    # 服务月首周一（跳过 28 日起的不完整首周）
    svc = _svc_start(cutoff.month, cutoff.year)
    ws0 = svc
    while ws0.weekday() != 0:
        ws0 += datetime.timedelta(days=1)

    def week_key(dt):
        if not hasattr(dt, "year"):
            return None
        if isinstance(dt, datetime.datetime):
            dt = dt.date()
        monday = dt - datetime.timedelta(days=dt.weekday())
        if monday < ws0:
            return None
        return monday.strftime("%Y-%m-%d")

    weeks = []
    d = ws0
    while d <= cutoff:
        if d.weekday() == 0:
            start, end = d, d + datetime.timedelta(days=6)
            weeks.append({"key": start.strftime("%Y-%m-%d"),
                          "label": "%s-%s" % (start.strftime("%m/%d"), end.strftime("%m/%d")),
                          "start": start.strftime("%Y-%m-%d"), "end": end.strftime("%Y-%m-%d")})
        d += datetime.timedelta(days=1)

    office = defaultdict(lambda: defaultdict(lambda: {"清洗保养": 0.0, "延保服务": 0.0, "清洁耗材": 0.0, "total": 0.0}))
    net = defaultdict(lambda: defaultdict(lambda: {"office": "—", "清洗保养": 0.0, "延保服务": 0.0, "清洁耗材": 0.0, "total": 0.0}))
    m2o = {}

    def add_map(name, off):
        if name and off and name not in m2o:
            m2o[name] = off

    def g(r, c):
        return r[c] if c < len(r) else None

    # 清洗保养 / 延保服务 : CSM服务项目
    ws = find_sheet("CSM服务项目")
    for r in list(ws.iter_rows(values_only=True))[2:]:
        cls = g(r, 16)
        if cls not in ("清洗保养", "延保服务"):
            continue
        wk = week_key(g(r, 7))
        if not wk:
            continue
        off = g(r, 1) or "—"
        nm = g(r, 3) or "未命名网点"
        amt = _num(g(r, 24))
        office[wk][off][cls] += amt
        office[wk][off]["total"] += amt
        net[wk][nm]["office"] = off
        net[wk][nm][cls] += amt
        net[wk][nm]["total"] += amt
        add_map(nm, off)
    # 清洁耗材 : CSM配件
    ws = find_sheet("CSM配件")
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if g(r, 17) != "清洁耗材":
            continue
        wk = week_key(g(r, 36))
        if not wk:
            continue
        off = g(r, 0) or "—"
        nm = g(r, 3) or "未命名网点"
        amt = _num(g(r, 27))
        office[wk][off]["清洁耗材"] += amt
        office[wk][off]["total"] += amt
        if net[wk][nm]["office"] in ("—", None):
            net[wk][nm]["office"] = off
        net[wk][nm]["清洁耗材"] += amt
        net[wk][nm]["total"] += amt
        add_map(nm, off)
    # 清洁耗材 : WMS网点买断配件明细
    ws = find_sheet("WMS网点买断配件明细")
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if g(r, 10) != "清洁耗材":
            continue
        wk = week_key(g(r, 11))
        if not wk:
            continue
        nm = g(r, 3) or "未命名网点"
        off = m2o.get(nm, "—")
        amt = _num(g(r, 8))
        office[wk][off]["清洁耗材"] += amt
        office[wk][off]["total"] += amt
        if net[wk][nm]["office"] in ("—", None):
            net[wk][nm]["office"] = off
        net[wk][nm]["清洁耗材"] += amt
        net[wk][nm]["total"] += amt

    office_out = {wk: {o: dict(v) for o, v in office[wk].items()} for wk in office}
    net_out = {wk: {n: dict(v) for n, v in net[wk].items()} for wk in net}
    return {"meta": {
                "sample": False,
                "note": "真实周数据：清洗保养/延保服务取自CSM服务项目(录入完成时间,上缴金额)；"
                        "清洁耗材取自CSM配件(录入完成时间,上缴金额)+WMS网点买断配件明细(发货时间,实收/上缴金额)。"
                        "首周(服务月28日起)已跳过。网点维度为源表维修商/经销部名。",
                "截止": cutoff.strftime("%Y-%m-%d")},
            "weeks": weeks, "office": office_out, "net": net_out}


# ---------------- 部署（GitHub Git Data API，内联，无需 _gh_push.py）----------------
API = "https://api.github.com"
OWNER, REPO, BRANCH = "mouren2580", "fangtai-dashboard", "main"
PUBLISH_DIR = "publish_gh"


def _req(method, url, token, payload=None, retries=4):
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    last = None
    for i in range(retries):
        r = urllib.request.Request(url, data=data, method=method)
        r.add_header("Authorization", "token " + token)
        r.add_header("Accept", "application/vnd.github+json")
        r.add_header("Content-Type", "application/json")
        r.add_header("User-Agent", "wb-sync")
        try:
            ctx = ssl.create_default_context()
            with urllib.request.urlopen(r, timeout=120, context=ctx) as resp:
                body = resp.read().decode("utf-8")
                return resp.status, (json.loads(body) if body.strip() else {})
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", "replace")
            try:
                j = json.loads(body)
            except Exception:
                j = {"message": body[:400]}
            if 400 <= e.code < 500 and e.code not in (403, 429):
                return e.code, j
            last = (e.code, j)
        except Exception as e:
            last = (0, {"message": repr(e)})
        wait = 2 * (i + 1)
        print("    retry %d/%d after %ds (%s)" % (i + 1, retries, wait, last[1].get("message", "")[:80]))
        time.sleep(wait)
    return last


def deploy(token, message):
    base = "%s/repos/%s/%s" % (API, OWNER, REPO)
    # 准备 publish_gh 目录
    os.makedirs(PUBLISH_DIR, exist_ok=True)
    idx = os.path.join(PUBLISH_DIR, "index.html")
    if not os.path.exists(idx):
        raise SystemExit("发布失败：找不到 %s，请先确保 dashboard.html 已生成" % idx)
    open(os.path.join(PUBLISH_DIR, ".nojekyll"), "w").close()
    files = []
    for root, _d, names in os.walk(PUBLISH_DIR):
        for nm in names:
            p = os.path.join(root, nm)
            rel = os.path.relpath(p, PUBLISH_DIR).replace("\\", "/")
            files.append((rel, p))
    files.sort()
    print("发布文件: %d" % len(files))
    tree = []
    for rel, p in files:
        raw = open(p, "rb").read()
        code, j = _req("POST", base + "/git/blobs", token,
                       {"content": base64.b64encode(raw).decode("ascii"), "encoding": "base64"})
        if code not in (200, 201):
            raise SystemExit("BLOB FAIL %s -> %s %s" % (rel, code, j.get("message")))
        tree.append({"path": rel, "mode": "100644", "type": "blob", "sha": j["sha"]})
    parents, base_tree = [], None
    code, j = _req("GET", base + "/git/ref/heads/" + BRANCH, token, retries=2)
    if code == 200 and "object" in j:
        parents = [j["object"]["sha"]]
        c2, jc = _req("GET", base + "/git/commits/" + parents[0], token, retries=2)
        if c2 == 200:
            base_tree = jc["tree"]["sha"]
    payload = {"tree": tree}
    if base_tree:
        payload["base_tree"] = base_tree
    code, j = _req("POST", base + "/git/trees", token, payload)
    if code not in (200, 201):
        raise SystemExit("TREE FAIL %s %s" % (code, j.get("message")))
    code, j = _req("POST", base + "/git/commits", token,
                   {"message": message, "tree": j["sha"], "parents": parents})
    if code not in (200, 201):
        raise SystemExit("COMMIT FAIL %s %s" % (code, j.get("message")))
    commit_sha = j["sha"]
    if parents:
        code, j = _req("PATCH", base + "/git/refs/heads/" + BRANCH, token, {"sha": commit_sha, "force": False})
    else:
        code, j = _req("POST", base + "/git/refs", token, {"ref": "refs/heads/" + BRANCH, "sha": commit_sha})
    if code not in (200, 201):
        raise SystemExit("REF FAIL %s %s" % (code, j.get("message")))
    print("OK  发布成功 -> https://github.com/%s/%s (commit %s)" % (OWNER, REPO, commit_sha[:10]))


# ---------------- 主流程 ----------------
def main():
    ap = argparse.ArgumentParser(description="方太西北服务产品看板 一键同步")
    ap.add_argument("--excel", required=True, help="最新月报 Excel 路径")
    ap.add_argument("--date", required=True, help="数据截止日 YYYY-MM-DD（更新日的前一天）")
    ap.add_argument("--html", default=os.path.join(HERE, "dashboard.html"), help="要更新的 dashboard.html（默认同目录）")
    ap.add_argument("--token", default=None, help="GitHub 个人令牌（也可放 .gh_token 文件）")
    ap.add_argument("--token-file", default=".gh_token", help="存放令牌的文件（默认 .gh_token）")
    ap.add_argument("--bump-stamp", action="store_true", help="推进「数据更新时间」戳（仅确认拿到新数据对外更新时用）")
    ap.add_argument("--no-deploy", action="store_true", help="只刷新本地文件，不发布到 GitHub")
    a = ap.parse_args()

    y, m, d = map(int, a.date.split("-"))
    as_of = datetime.date(y, m, d)
    cutoff = as_of  # 周数据也截到同一天

    if not os.path.exists(a.excel):
        raise SystemExit("找不到 Excel：%s" % a.excel)

    # 自动定位 dashboard.html：同目录 → 上级目录（适配 sync-kit 子目录场景）
    if not os.path.exists(a.html):
        for cand in [os.path.join(HERE, "dashboard.html"),
                     os.path.join(HERE, "..", "dashboard.html"),
                     os.path.join(HERE, "..", "..", "dashboard.html")]:
            if os.path.exists(cand):
                a.html = cand
                break
    if not os.path.exists(a.html):
        raise SystemExit("找不到 dashboard.html：默认在同目录或上级目录（sync-kit 场景），可用 --html 指定")
    html = open(a.html, encoding="utf-8").read()

    # 1) 主数据 + 服务工程师
    data = build(a.excel, as_of)
    old_stamp = None
    mm = re.search(r"(?:const|let) DASHBOARD_DATA = (\{.*?\});", html, re.S)
    if mm:
        try:
            old_stamp = (json.loads(mm.group(1)).get("meta", {}) or {}).get("_stamp")
        except Exception:
            old_stamp = None
    if a.bump_stamp:
        stamp = datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")
    else:
        stamp = old_stamp or datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")
    data["meta"]["_stamp"] = stamp
    tech, split = build_tech(a.excel)
    html = replace_var(html, "DASHBOARD_DATA", data)
    html = replace_var(html, "TECH_DATA", tech)

    # 2) 延保
    ext = gen_extend(a.excel, as_of)
    html = replace_var(html, "EXTEND_WARRANTY_DATA", ext)

    # 3) 周数据
    wk = gen_weekly(a.excel, cutoff)
    html = replace_var(html, "WEEKLY_DATA", wk)

    open(a.html, "w", encoding="utf-8").write(html)

    # 汇总打印
    actual = sum(n["合计"] for n in data["网点"])
    t = data["time"]
    print("✅ 已刷新本地:", os.path.abspath(a.html))
    print("   截止日=%s 周期=%s 已完成%d/%d天 (%.1f%%) 剩余%d天"
          % (as_of, t["周期标签"], t["日"], t["月度天数"], t["日"] / t["月度天数"] * 100, t["月度天数"] - t["日"]))
    print("   网点=%d 总实际=¥%s" % (len(data["网点"]), format(round(actual, 2), ",")))
    print("   工程师=%d 工单消耗=¥%s 网点买断=¥%s 报表总=¥%s"
          % (tech["count"], format(tech["total"], ","), format(tech["meta"]["网点买断"], ","), format(tech["meta"]["报表总合计"], ",")))
    print("   延保合计: 一年 ¥%.2f / 二年 ¥%.2f / 三年 ¥%.2f"
          % (ext["total"]["y1"]["amt"], ext["total"]["y2"]["amt"], ext["total"]["y3"]["amt"]))
    wk_tot = sum(v["total"] for wk in wk["office"].values() for v in wk.values())
    print("   周数据: %d 周, 三大爆品合计 ¥%.2f" % (len(wk["weeks"]), wk_tot))
    print("   数据来源拆分:", {k: round(v, 2) for k, v in split.items()})
    print("   _stamp =", stamp)

    # 4) 部署
    if a.no_deploy:
        print("⏭️  跳过发布（--no-deploy）。需要上线时去掉该参数并加 --token。")
        return
    token = a.token or (open(a.token_file, encoding="utf-8").read().strip() if os.path.exists(a.token_file) else None)
    if not token:
        raise SystemExit("未提供令牌：加 --token ghp_xxx 或确保 %s 存在" % a.token_file)
    # 同步 publish_gh/index.html
    os.makedirs(PUBLISH_DIR, exist_ok=True)
    import shutil
    shutil.copyfile(a.html, os.path.join(PUBLISH_DIR, "index.html"))
    msg = "同步{d}数据(截止{date})".format(d=as_of.strftime("%Y-%m"), date=a.date)
    deploy(token, msg)


if __name__ == "__main__":
    main()
