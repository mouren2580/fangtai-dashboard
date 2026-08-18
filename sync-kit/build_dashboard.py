#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
方太 西北大区 服务产品达成看板 - 构建脚本
读取原始月报 Excel（多 sheet 交易明细），提取「服务产品收入统计」表
（天然是 服务网点 × 销售分类(品项) 矩阵），聚合成 办事处/服务网点/品项 结构，
注入 HTML 模板，产出单文件离线看板 dashboard.html。

用法:
  python build_dashboard.py <月报.xlsx> [输出dashboard.html] [--date YYYY-MM-DD]

依赖: openpyxl
"""
import os, sys, json, argparse, datetime
from collections import defaultdict

# Windows 控制台默认 GBK，输出中文/emoji 会 UnicodeEncodeError，强制 UTF-8
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "..", "assets", "dashboard_template.html")
SHEETJS = os.path.join(HERE, "..", "assets", "sheetjs.min.js")

def f(v):
    if v is None:
        return 0.0
    try:
        return round(float(v), 2)
    except Exception:
        return 0.0

def find_sheet(wb):
    """定位服务产品收入统计表：表名含 '服务产品收入统计' 或首行含 '销售分类'。"""
    for ws in wb.worksheets:
        if "服务产品收入统计" in ws.title:
            return ws
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        for r in rows:
            if r and any("数据来源" == (c or "").strip() and "合计" in [str(x) for x in r] for c in r):
                return ws
    raise RuntimeError("未找到『服务产品收入统计』表")

# ---------- 服务月周期表（2026 用户口径，2026-08-15 确认） ----------
# 通用规则：上月28日 → 当月27日。
# 2026 特例：2月止于 2.25、3月起于 2.26（其余月份同通用规则，全年 365 天无缝衔接 2025.12.28→2026.12.27）。
# 用 (年,月) 作 key，值为 (start_iso, end_iso)；缺失则退回通用规则。
SERVICE_PERIOD_OVERRIDE = {
    (2026, 1): ("2025-12-28", "2026-01-27"),
    (2026, 2): ("2026-01-28", "2026-02-25"),
    (2026, 3): ("2026-02-26", "2026-03-27"),
    (2026, 4): ("2026-03-28", "2026-04-27"),
}
def service_period(y, m):
    key = (y, m)
    if key in SERVICE_PERIOD_OVERRIDE:
        s, e = SERVICE_PERIOD_OVERRIDE[key]
        return datetime.date.fromisoformat(s), datetime.date.fromisoformat(e)
    ps = datetime.date(y - 1, 12, 28) if m == 1 else datetime.date(y, m - 1, 28)
    pe = datetime.date(y, m, 27)
    return ps, pe

def build(excel_path, as_of, period_start=None, period_end=None):
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = find_sheet(wb)
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    # 定位表头行（含 '数据来源' 与 '合计'）
    hdr_idx = None
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if "数据来源" in cells and "合计" in cells:
            hdr_idx = i
            break
    if hdr_idx is None:
        raise RuntimeError("表头未找到 数据来源/合计 列")
    hdr = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
    src_i = hdr.index("数据来源")
    tot_i = hdr.index("合计")
    item_cols = hdr[src_i + 1: tot_i]          # 品项（销售分类）
    # 关键维: 大区/服务中心/办事处/服务网点
    di = hdr.index("大区") if "大区" in hdr else 0
    si = hdr.index("服务中心") if "服务中心" in hdr else 1
    oi = hdr.index("办事处") if "办事处" in hdr else 2
    ni = hdr.index("服务网点") if "服务网点" in hdr else 3

    nets = {}
    for r in rows[hdr_idx + 1:]:
        if r[di] is None and r[oi] is None:
            continue
        办事处 = (r[oi] or "").strip()
        服务中心 = (r[si] or "").strip()
        服务网点 = (r[ni] or "").strip()
        来源 = (r[src_i] or "").strip()
        if not 服务网点:
            continue
        key = (办事处, 服务中心, 服务网点)
        d = nets.setdefault(key, {"品项": {it: 0.0 for it in item_cols}, "合计": 0.0, "来源": set()})
        for it in item_cols:
            d["品项"][it] += f(r[hdr.index(it)])
        d["合计"] += f(r[tot_i])
        if 来源:
            d["来源"].add(来源)

    # 时间进度（方太服务月惯例：上月28日 → 本月27日；可用 --start/--end 覆盖）
    y, m = as_of.year, as_of.month
    if period_start is None or period_end is None:
        ps, pe = service_period(y, m)
        period_start, period_end = ps, pe
    total_days = (period_end - period_start).days + 1     # 周期总天数（含首尾）
    elapsed = (as_of - period_start).days + 1             # 数据截止日（含当天）在周期中的天数
    elapsed = max(1, min(elapsed, total_days))
    period_label = f"{period_start.month}.{period_start.day}–{period_end.month}.{period_end.day}（{total_days}天）"
    week = (elapsed - 1) // 7 + 1
    weeks = (total_days - 1) // 7 + 1

    out = {
        "meta": {"大区": "西北大区部", "月份": f"{y}-{m:02d}", "品项": item_cols,
                 "来源表": "服务产品收入统计",
                 # 统计截止日：供前端 DEFAULT_ASOF 读取，避免模板硬编码导致换月时间进度算错
                 "统计截止日": as_of.isoformat()},
        "网点": [{"办事处": k[0], "服务中心": k[1], "服务网点": k[2],
                  "合计": round(d["合计"], 2),
                  "品项": {it: round(d["品项"][it], 2) for it in item_cols},
                  "来源": sorted(d["来源"])} for k, d in nets.items()],
        "time": {"年": y, "月": m, "日": elapsed, "月度天数": total_days,
                 "周": week, "总周数": weeks,
                 "周期开始": period_start.isoformat(), "周期结束": period_end.isoformat(),
                 "周期标签": period_label},
    }
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", help="月报 Excel 路径")
    ap.add_argument("out", nargs="?", default="dashboard.html")
    ap.add_argument("--date", default=None, help="数据截止日 YYYY-MM-DD，默认今天")
    ap.add_argument("--start", default=None, help="统计周期开始 YYYY-MM-DD（默认服务月：上月28日）")
    ap.add_argument("--end", default=None, help="统计周期结束 YYYY-MM-DD（默认服务月：本月27日）")
    args = ap.parse_args()
    as_of = datetime.date.today()
    if args.date:
        y, m, d = map(int, args.date.split("-"))
        as_of = datetime.date(y, m, d)
    ps = pe = None
    if args.start and args.end:
        sy, sm, sd = map(int, args.start.split("-"))
        ey, em, ed = map(int, args.end.split("-"))
        ps, pe = datetime.date(sy, sm, sd), datetime.date(ey, em, ed)
    elif args.start or args.end:
        print("⚠ 仅提供了 --start/--end 之一，已忽略并使用服务月周期。")
    data = build(args.excel, as_of, ps, pe)
    tpl = open(TEMPLATE, encoding="utf-8").read()
    assert "__DATA__" in tpl, "模板缺少 __DATA__ 占位符"
    assert "__SHEETJS__" in tpl, "模板缺少 __SHEETJS__ 占位符"
    sheetjs = open(SHEETJS, encoding="utf-8").read() if os.path.exists(SHEETJS) else ""
    html = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    html = html.replace("__SHEETJS__", sheetjs)
    open(args.out, "w", encoding="utf-8").write(html)
    t = data["time"]
    print(f"✅ 已生成 {args.out}")
    print(f"   月份={data['meta']['月份']} 网点={len(data['网点'])} 品项={len(data['meta']['品项'])} "
          f"总实际=¥{sum(n['合计'] for n in data['网点']):,.0f}")
    print(f"   统计周期={t['周期标签']} 时间进度={t['日']}/{t['月度天数']}="
          f"{(t['日']/t['月度天数']*100):.1f}% 第{t['周']}周/共{t['总周数']}周")

if __name__ == "__main__":
    main()
